import pandas as pd
from openpyxl.chart import BarChart, Reference
import openpyxl
import os

PATH = r"pandas_chart3556numpy.xlsx"
SHEETNAME = "benben_sheet43"


def add_barchart_sheet_to_excel(excel_path, name, df):
	"""
	*add chart sheet using "df" data, to existing / new workbook (excel file)
	*all df data will be inserted to sheet, but only first 2 columns will be used for the chart

	:str: excel_path: path for the relevant excel
	:str: name: name for the sheet and chart
	:pandas.Dataframe: df: the data to insert to sheet and base the chart on
	"""
	# create excel if doesn't exist
	remove_default_sheet = False
	if not os.path.isfile(excel_path):
		wb = openpyxl.Workbook().save(excel_path)
		remove_default_sheet = True

	# init workbook
	workbook = openpyxl.load_workbook(excel_path)
	if name in workbook.sheetnames:
		raise Exception('"{0}" sheet already exists in "{1}"'.format(name, excel_path))
	writer = pd.ExcelWriter(excel_path, engine='openpyxl')
	writer.book = workbook

	# create sheet and add df to it 
	df.to_excel(writer, sheet_name=name)
	writer.save()
	writer.close()

	# set chart attributes
	chart = BarChart()
	chart.type = "col"
	chart.style = 10
	chart.title = name
	chart.x_axis.title = df.columns[0]
	chart.y_axis.title = df.columns[1]

	# create chart
	worksheet = workbook[name]
	# categories == x-axis, data == y-axis
	categories = Reference(worksheet, min_col=2, min_row=2, max_row=len(df) + 2) 
	data = Reference(worksheet, min_col = 3, min_row = 1, max_row=len(df) + 2) 
	chart.add_data(data, titles_from_data=True)
	chart.set_categories(categories)
	chart.shape = 4
	worksheet.add_chart(chart, "A{0}".format(len(df) + 5))

	# remove default sheet if created the excel in current function iteration
	if remove_default_sheet:
		del workbook["Sheet"]

	workbook.save(excel_path)
	workbook.close()


df = pd.DataFrame({'Mahzor': [177, 178, 179, 180], 'Score': [6.5, 6.6,6.7, 6.8], 'Number of Answers': [3,2,1,2]})
add_barchart_sheet_to_excel(PATH, SHEETNAME, df)

