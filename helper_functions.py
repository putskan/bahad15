from email.mime.text import MIMEText
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email import encoders
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from flask_init import *
import smtplib, ssl
import codecs
import pandas as pd
import re
import openpyxl
import shutil
import os
import constants


def send_reports_email(port, smtp_server, sender_email, receiver_email, password, email_content):
    """
    """
    context = ssl.create_default_context()
    with smtplib.SMTP(smtp_server, port) as server:
        server.ehlo()  # Can be omitted
        server.starttls(context=context)
        server.ehlo()  # Can be omitted
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, email_content)
        server.quit()


def archive_reports():
    """
    """
    report_paths = get_report_paths()
    for path in report_paths:
        basedir, filename = os.path.split(path)
        new_path = os.path.join(basedir, 'Archive/', filename)
        shutil.move(path, new_path)


def get_report_paths():
    """
    """
    report_paths = list(filter(lambda file_path: os.path.isfile(file_path), map(lambda filename: os.path.join(constants.REPORTS_PATH, filename), os.listdir(constants.REPORTS_PATH))))
    return report_paths


def generate_basic_report(table_name, form_html_filename):
    """
    """
    # eg. [{hebrew_description: hebrewtext, english_name: englishtext}, {}, ...]
    medical_full_tbl_df = pd.read_sql_query("SELECT * FROM {0}".format(table_name), db.engine)
    input_hebrew_descs_to_names = english_input_name_to_hebrew_desc(os.path.join(constants.WORKING_DIR, 'templates/{0}'.format(form_html_filename)))
    #  will be used for replacement of df column names from english to hebrew
    df_col_name_replacement_dict = {}
    for hebrew_english_link in input_hebrew_descs_to_names:
        df_col_name_replacement_dict[hebrew_english_link["english_name"]] = hebrew_english_link["hebrew_description"]

    medical_full_tbl_df = medical_full_tbl_df.rename(columns=df_col_name_replacement_dict)
    medical_full_tbl_df = medical_full_tbl_df.replace(["yes", "no"], ["כן", "לא"])
    medical_full_tbl_df = medical_full_tbl_df.drop("_id", axis=1)
    return medical_full_tbl_df
    
    
def english_input_name_to_hebrew_desc(html_path):
    """
    link between <input> / <textarea> "name" attribute value, to its hebrew description, in an html document

    :str: html_path: path to an html file
    return:
        :list: [{hebrew_description: hebrewtext, english_name: englishtext}, {}, ...]
    """
    with codecs.open(html_path, encoding='utf-8') as f:
        doc_content = f.read()

    hebrew_to_input_name = re.findall(constants.LINK_ENGLISH_INPUT_NAME_TO_HEBREW_DESCRIPTION_REGEX, doc_content)
    # eliminate regex empty values and convert to list. eg, from:
    # ('מספר אישי', 'private_number', '', '', '', '')
    # to:
    # ['מספר אישי', 'private_number']
    hebrew_to_input_name = list(list((v for v in re_match if v != '')) for re_match in hebrew_to_input_name)
    # convert inner-lists to dicts. eg: {'hebrew_description': 'מספר אישי', 'english_name': 'private_number'}
    hebrew_to_input_name = list(dict(hebrew_description = hebrew_input_pair[0], english_name = hebrew_input_pair[1]) for hebrew_input_pair in hebrew_to_input_name)
    return hebrew_to_input_name


def build_reports_email_content():
    """
    build and return email content
    """
    # add meta
    msg = MIMEMultipart('alternative')
    msg.set_charset('utf8')
    msg['FROM'] = constants.SENDER_EMAIL
    msg['To'] = constants.RECIEVER_EMAIL
    
    message_subject = 'Kamanim Report - {current_date}'.format(current_date=datetime.now().strftime("%m/%d/%Y %H:%M"))
    msg['Subject'] = Header(message_subject.encode('utf-8'), 'UTF-8').encode()
    # add message body / content
    message_body = 'מצ"ב הדו"חות הרלוונטיים.'
    msg.attach(MIMEText(message_body.encode('utf-8'), 'html', 'UTF-8'))

    # add report files
    report_paths = get_report_paths()
    for report_path in report_paths:
        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(report_path, "rb").read())
        encoders.encode_base64(part)
        path, filename = os.path.split(report_path)
        part.add_header('Content-Disposition', 'attachment; filename={filename}'.format(filename=filename))
        msg.attach(part)

    return msg


def string_converter(my_str):
    """
    convert strings recieved via POST to python types (datetime, boolean, etc)
    """
    if my_str == '':
        return None
    # don't use bool, because sqlite replaces them with 1,0 (which are later harder to manipluate)
    if my_str.upper() == "TRUE":
        return "yes"
    if my_str.upper() == "FALSE":
        return "no"
    if re.findall(constants.DATE_REGEX, my_str):
        return datetime.strptime(my_str, "%Y-%m-%d")
    return my_str


def add_barchart_sheet_to_excel(excel_path, name, df):
    """
    *add chart sheet using "df" data, to existing / new workbook (excel file)
    *all df data will be inserted to sheet, but only first 2 columns will be used for the chart

    :str: excel_path: path for the relevant excel
    :str: name: name for the sheet and chart
    :pandas.Dataframe: df: the data to insert to sheet and base the chart on
    """
    # create excel if doesn't exist
    remove_default_sheet = False
    if not os.path.isfile(excel_path):
        wb = openpyxl.Workbook().save(excel_path)
        remove_default_sheet = True

    # init workbook
    workbook = openpyxl.load_workbook(excel_path)
    if name in workbook.sheetnames:
        raise Exception('"{0}" sheet already exists in "{1}"'.format(name, excel_path))
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = workbook

    # create sheet and add df to it 
    df.to_excel(writer, sheet_name=name)
    writer.save()
    writer.close()

    # set chart attributes
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = name
    chart.x_axis.title = df.columns[0]
    chart.y_axis.title = df.columns[1]

    # create chart
    worksheet = workbook[name]
    # categories == x-axis, data == y-axis
    categories = Reference(worksheet, min_col=2, min_row=2, max_row=len(df) + 2) 
    data = Reference(worksheet, min_col = 3, min_row = 1, max_row=len(df) + 2) 
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.shape = 4
    worksheet.add_chart(chart, "A{0}".format(len(df) + 5))

    # remove default sheet if created the excel in current function iteration
    if remove_default_sheet:
        del workbook["Sheet"]

    workbook.save(excel_path)
    workbook.close()

